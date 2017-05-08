def xlsxwriter_ex(excel_file):
    print("[==== xlsxwriter example ====]") 
        
    # Limitations with writing to Excel using xlsxwriter:
    # -Cannot sort a column or at least its not recommended:
    
    import csv
    import xlsxwriter
    
    cars = []
    
    # Populate the data we'll write to Excel from in.csv
    with open('in.csv') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=",")
        for row in reader:
            car = {}
            for key in row.keys():
                car[key] = row[key]
            cars.append(car)

    # Create our excel objects
    wb = xlsxwriter.Workbook(excel_file)
    ws = wb.add_worksheet("Cars")
    
    row = 0
    col = 0
    
    header_format = wb.add_format({"bold": True,
                                   "bg_color": "#1e3966",
                                   "border": 1,
                                   "align": "center",
                                   "font_color": "white"})
    # Write the header using the fields from the first car in the list
    for column in cars[0].keys():
        ws.write_string(row, col, column, header_format)
        col += 1
    
    # Creates a border around each cell so it looks like a grid
    data_format = wb.add_format({"border": 1})
    
    # Everything is written using a row/column index.  We must increment the
    # row index for each line and the column index for each value
    for car in cars:
        row += 1 
        col = 0
        for key in car.keys():
            value = car[key]
            ws.write_string(row, col, value, data_format)
            col += 1
    
    # I could not find a way to set a filter only on columns that are populated.
    # An automated way of doing this is getting the # of columns and translating
    # that to the appropriate column letter as the end range.
    ws.autofilter("A1:C1")
    # Freeze top row and left column
    ws.freeze_panes(1, 1)
    wb.close()
    
def pyexcel_ex(input_file):
    print("[==== pyexcel example ====]") 
    import pyexcel as pe
    records = pe.iget_records(file_name=input_file)

    # Using pyexcel is preferable because it will automatically create a dictionary
    # of each row that uses the 1st row column values as the keys
    for record in records:
        print("Make={} Model={} Year={}".format(record['Make'], 
                                                record['Model'], 
                                                record['Year']))

def openpyxl_ex(input_file):
    print("[==== openpyxl example ====]") 
    from openpyxl import load_workbook
    wb = load_workbook(input_file, read_only=True)
    print("Sheet names={}".format(wb.get_sheet_names()))

    ws = wb['Cars']

    for row in ws.rows:
        for cell in row:
            print(cell.value)
            
    # As far as I know, openpyxl cannot access values using a dictionary 
    # structure like pyexcel
    print("Cell A1={}".format(ws['A1'].value))

excel_file = "cars.xlsx"
        
xlsxwriter_ex(excel_file)
pyexcel_ex(excel_file)
openpyxl_ex(excel_file)
